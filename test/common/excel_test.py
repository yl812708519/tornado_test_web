#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys

from openpyxl import load_workbook

from test.common.pinyin import Pinyin


reload(sys)
sys.setdefaultencoding('utf-8')
def get_customers_from_excel():
    print_flag = False
    print "开始读取文件……"
    wb2 = load_workbook('/home/wangshubin/dev/python/workspaces/inc-eking-web/test/common/crm_customers.xlsx')
    ws =  wb2.get_active_sheet()
    customers = []
    for i in range(2,1438):

        # 客户名称 法务人员 所属子公司 服务项目 录入时间 客户联系人 联系人电话 客户地址 组织机构代码/身份证 联系人QQ 联系人邮箱
        customer = dict()
        contracts = []
        customer['name'] = ws.cell(row = i, column = 1).value
        customer['dealer'] = ws.cell(row = i, column = 2).value
        customer['company'] = ws.cell(row = i, column = 3).value
        customer['service_case'] = ws.cell(row = i, column = 4).value #合同的信息
        customer['entry_time'] = ws.cell(row = i, column = 5).value #合同的录入时间
        customer['address'] = ws.cell(row = i, column = 8).value
        customer['certificate_number'] = ws.cell(row = i, column = 9).value

        if not customer['name'] or customer['name'] =="/":
            if print_flag:
                print i,' 行客户名称不存在或为“/”被忽略'
            continue
        if customer['name'].find("/")>=0:
            if print_flag:
                print i,' 行客户名称存在“/”被忽略：',customer['name']
            continue

        if not customer['certificate_number']:
            customer['certificate_number']=""
        if not customer['address']:
            customer['address'] = ""



        #处理联系人
        contract_name = ws.cell(row = i, column = 6).value
        contract_mobile = ws.cell(row = i, column = 7).value
        contract_qq = ws.cell(row = i, column = 10).value
        contract_email = ws.cell(row = i, column = 11).value

        contract_mobiles = []
        contract_qqs = []
        contract_emails = []
        if contract_mobile:
            contract_mobile = str(contract_mobile)
            contract_mobile = contract_mobile.replace(" ","")
            if contract_mobile.find("、")>=0:
                if print_flag:
                    print i," 行电话是使用 、 分割"
                contract_mobiles = contract_mobile.split("、")
            else:
                contract_mobiles = contract_mobile.split("/")
        if contract_qq:
            contract_qq = str(contract_qq)
            contract_qq = contract_qq.replace(" ","")
            if contract_mobile.find("、")>=0:
                if print_flag:
                    print i," 行QQ是使用 、 分割"
                contract_qqs = contract_qq.split("、")
            else:
                contract_qqs = contract_qq.split("/")
        if contract_email:
            contract_email = str(contract_email)
            contract_email = contract_email.replace(" ","")
            if contract_mobile.find("、")>=0:
                if print_flag:
                    print i," 行邮箱是使用 、 分割"
                contract_emails = contract_email.split("、")
            else:
                contract_emails = contract_email.split("/")


        if not contract_name:
            contract_name = customer['name']
        contract_names = contract_name.split("/")

        for ii in range(len(contract_names)):
            contract = dict()
            contract["name"] = contract_names[ii]
            if contract_mobiles and ii in range(len(contract_mobiles)):
                contract["mobile"]=contract_mobiles[ii]
            if contract_qqs and ii in range(len(contract_qqs)):
                contract["qq"]=contract_qqs[ii]
            if contract_emails and ii in range(len(contract_emails)):
                contract["email"]=contract_emails[ii]
            # print contract
            contracts.append(contract)
        customer["contracts"] = contracts
        #处理法务
        if not customer['dealer']:
            if print_flag:
                print i, " 行法务为空,已默认为客户名称"
            customer['dealer'] = customer['name']
        dealer_name = customer['dealer']
        dealers = get_dealers(dealer_name)
        customer["dealers"] = dealers
        # print dealers
        customers.append(customer)
    # end of for
    print "文件读取结束，开始写入数据库……"
    return customers
def get_dealers(dealer_name):
    """
    保存法务
    :param dealer_name:
    :return: 法务list
    """
    pinyin = Pinyin()
    dealers = []
    if dealer_name.find("、")>=0:
        dealer_names = dealer_name.split("、")
    else:
        dealer_names = dealer_name.split("/")
    for i in range(len(dealer_names)):
        deal_name = dealer_names[i]
        dealer = dict()
        dealer["name"] = deal_name
        dealer["account"] = pinyin.get_pinyin(deal_name,"")+"@eking.mobi"
        dealers.append(dealer)
    return dealers

if __name__ == '__main__':
    # print Settings.STATIC_SITE_ROOT + Settings.STATIC_URL_PREFIX + 'img/avatar-default.gif'
    # print Settings.STATIC_URL_PREFIX + 'img/avatar-default.gif'
    get_customers_from_excel()